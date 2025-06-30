import json
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import requests

CF_RATING_COLOR = {
    'unrated': '#000000',
    'headquarters': '#000000',
    'newbie': '#808080',
    'pupil': '#008000',
    'specialist': '#03a89e',
    'expert': '#0000ff',
    'candidate master': '#aa00aa',
    'master': '#ff8c00',
    'grandmaster': '#ff0000',
    'international master': '#ff0000',
}

NC_RATING_COLOR = {
    'rate-score6': '#ff8800',
    'rate-score5': '#ffd700',
    'rate-score4': '#25bb9b',
    'rate-score3': '#5ea1f4',
    'rate-score2': '#c177e7',
    'rate-score1': '#b4b4b4',
}


wb = load_workbook('info/id.xlsx')

info: dict[int, list[dict]] = {}
cf_ids: list[str] = []
for ws in wb:
    users = info[ws.title] = []
    for row in map(lambda row: tuple(map(lambda v: str(v).strip() if v else v, row)), ws.iter_rows(2, values_only=True)):
        if not row[0]:
            break

        if len(row) >= 5 and row[4]:
            continue

        users.append({
            'name': row[0],
            'cf': row[1],
            'nc': row[2],
            'atc': row[3],
        })
        if row[1] is not None:
            cf_ids.append(row[1])

wb.close()


# 由于 cf 获取 rating api 允许一次获取多个账号，因此先统一获取
cf_url = f"https://codeforces.com/api/user.info?checkHistoricHandles=false&handles={';'.join(cf_ids)}"
cf_response: dict = requests.get(cf_url).json()

if cf_response['status'] != 'OK':
    raise Exception(cf_response['comment'])

cf_res = {}
for user in cf_response['result']:
    rated = 'rating' in user
    cf_res[user['handle'].lower()] = {
        'id': user['handle'],
        'rating': user['rating'] if rated else 'unrated',
        'color': CF_RATING_COLOR[user['rank'] if rated else 'unrated']
    }


for users in info.values():
    for user in users:
        # Codeforces
        cf_id, user['cf'] = user['cf'], {}
        if cf_id is not None:
            user['cf'] = cf_res[cf_id.lower()]

        # 牛客
        nc_id, user['nc'] = user['nc'], {}
        if nc_id is not None:
            nc_url = f"https://ac.nowcoder.com/acm/contest/rating-index?searchUserName={nc_id}"
            nc_res = requests.get(nc_url).text

            bs = BeautifulSoup(nc_res, 'html.parser')
            tbody = bs.find('tbody')
            if tbody is not None:
                tr = bs.find('tbody').tr
                span = tr.find_all('td')[4].span
                user['nc'] = {
                    'id': nc_id,
                    'uid': tr['data-uid'],
                    'rating': int(span.get_text()),
                    'color': NC_RATING_COLOR[span['class'][0]]
                }

        # AtCoder
        atc_id, user['atc'] = user['atc'], {}
        if atc_id is not None:
            user['atc']['id'] = atc_id


DOCS = Path('docs/public')

DOCS.mkdir(exist_ok=True)
with open(DOCS / 'info.json', 'w') as f:
    json.dump(info, f)
